#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
File Server Audit Tool
================================================================================

Description:
    A comprehensive audit tool for Windows file servers designed to support
    migration activities. This script scans directory structures and generates
    detailed Excel reports containing file/folder inventory with NTFS permissions.

Features:
    - Recursive scanning of all folders and subfolders
    - Folder exclusion support with wildcard patterns
    - NTFS permission extraction (inherited vs explicit)
    - File metadata collection (size, extension, last modified date)
    - Owner information retrieval
    - Excel output with auto-filter and formatted headers
    - Error logging for access-denied scenarios
    - Progress tracking via console and log file

Output:
    - Excel file (.xlsx) with file/folder listing and permissions
    - Log file with detailed scan progress and errors

Requirements:
    pip install openpyxl pywin32

Usage:
    python file_server_audit.py "D:\\Data"
    python file_server_audit.py "D:\\Data" --exclude "D:\\Data\\Temp"
    python file_server_audit.py "D:\\Data" -x "*\\Archive" -x "*\\Backup"
    python file_server_audit.py "D:\\Data" --exclude-file exclusions.txt
    python file_server_audit.py "D:\\Data" -x "*\\Temp" -e exclusions.txt

Exclusion Options:
    -x, --exclude       Exclude a folder path pattern (can be used multiple times)
    -e, --exclude-file  Path to text file with exclusion patterns

Exclusion File Format:
    - One path pattern per line
    - Supports wildcards: * (any characters), ? (single character)
    - Case-insensitive matching
    - Lines starting with # are comments
    - Empty lines are ignored

    Example exclusions.txt:
        # Temporary folders
        D:\\Data\\Temp
        D:\\Data\\*\\tmp
        \\\\server\\share\\Archive*
        *\\backup_*

--------------------------------------------------------------------------------
Author:     Ishak Ahmad (ishak.ahmad@gmail.com)
Created:    2025
Version:    1.0.0
License:    Proprietary - All Rights Reserved

Copyright (c) 2025 Ishak Ahmad. All rights reserved.

This software is proprietary and confidential. Unauthorized copying,
distribution, modification, or use of this software, via any medium,
is strictly prohibited without prior written permission from the author.
================================================================================
"""

import os
import sys
import argparse
import logging
import fnmatch
from datetime import datetime
from pathlib import Path

try:
    import win32security
    import win32api
    import ntsecuritycon as con
except ImportError:
    print("ERROR: pywin32 is not installed. Please run: pip install pywin32")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is not installed. Please run: pip install openpyxl")
    sys.exit(1)


# Permission mapping for common access rights
PERMISSION_FLAGS = {
    con.FILE_READ_DATA: "Read",
    con.FILE_WRITE_DATA: "Write",
    con.FILE_APPEND_DATA: "Append",
    con.FILE_READ_EA: "ReadEA",
    con.FILE_WRITE_EA: "WriteEA",
    con.FILE_EXECUTE: "Execute",
    con.FILE_DELETE_CHILD: "DeleteChild",
    con.FILE_READ_ATTRIBUTES: "ReadAttr",
    con.FILE_WRITE_ATTRIBUTES: "WriteAttr",
    con.DELETE: "Delete",
    con.READ_CONTROL: "ReadControl",
    con.WRITE_DAC: "WriteDAC",
    con.WRITE_OWNER: "WriteOwner",
    con.SYNCHRONIZE: "Sync",
}

# Common permission combinations
GENERIC_PERMISSIONS = {
    con.FILE_ALL_ACCESS: "FullControl",
    con.FILE_GENERIC_READ: "Read",
    con.FILE_GENERIC_WRITE: "Write",
    con.FILE_GENERIC_EXECUTE: "Execute",
}


def load_exclusion_patterns(exclude_file, logger):
    """Load exclusion patterns from a text file.

    Args:
        exclude_file: Path to the exclusion file
        logger: Logger instance

    Returns:
        List of exclusion patterns (lowercase for case-insensitive matching)
    """
    patterns = []

    if not exclude_file:
        return patterns

    if not os.path.exists(exclude_file):
        logger.warning(f"Exclusion file not found: {exclude_file}")
        return patterns

    try:
        with open(exclude_file, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()

                # Skip empty lines and comments
                if not line or line.startswith('#'):
                    continue

                # Normalize path separators and convert to lowercase
                pattern = line.replace('/', '\\').lower()
                patterns.append(pattern)
                logger.debug(f"Loaded exclusion pattern: {pattern}")

        logger.info(f"Loaded {len(patterns)} exclusion pattern(s) from {exclude_file}")

    except Exception as e:
        logger.error(f"Error reading exclusion file: {str(e)}")

    return patterns


def is_path_excluded(path, exclusion_patterns):
    """Check if a path matches any exclusion pattern.

    Args:
        path: Full path to check
        exclusion_patterns: List of patterns (lowercase)

    Returns:
        True if path should be excluded, False otherwise
    """
    if not exclusion_patterns:
        return False

    # Normalize path for comparison (lowercase, consistent separators)
    normalized_path = path.replace('/', '\\').lower()

    for pattern in exclusion_patterns:
        # Check if pattern matches the path using fnmatch (supports * and ?)
        if fnmatch.fnmatch(normalized_path, pattern):
            return True

        # Also check if the path starts with the pattern (for exact directory matches)
        # This handles cases where pattern is "D:\Data\Temp" and path is "D:\Data\Temp\subfolder"
        pattern_normalized = pattern.rstrip('\\')
        if normalized_path == pattern_normalized or normalized_path.startswith(pattern_normalized + '\\'):
            return True

    return False


def setup_logging(output_dir):
    """Setup logging to both console and file."""
    log_file = output_dir / f"audit_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    # Create logger
    logger = logging.getLogger('FileAudit')
    logger.setLevel(logging.DEBUG)

    # Console handler - INFO level
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_format = logging.Formatter('%(asctime)s - %(message)s', datefmt='%H:%M:%S')
    console_handler.setFormatter(console_format)

    # File handler - DEBUG level
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_format = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(file_format)

    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    return logger


def format_size(size_bytes):
    """Convert bytes to human readable format."""
    if size_bytes == 0:
        return "0 B"

    units = ['B', 'KB', 'MB', 'GB', 'TB']
    unit_index = 0
    size = float(size_bytes)

    while size >= 1024 and unit_index < len(units) - 1:
        size /= 1024
        unit_index += 1

    return f"{size:.2f} {units[unit_index]}"


def get_access_mask_string(access_mask):
    """Convert access mask to readable permission string."""
    # Check for full control first
    if access_mask & con.FILE_ALL_ACCESS == con.FILE_ALL_ACCESS:
        return "FullControl"

    # Check for generic permissions
    permissions = []

    # Check common combinations
    if access_mask & 0x1F01FF == 0x1F01FF:  # Full Control
        return "FullControl"
    if access_mask & 0x1301BF == 0x1301BF:  # Modify
        return "Modify"
    if access_mask & 0x1200A9 == 0x1200A9:  # ReadAndExecute
        permissions.append("ReadAndExecute")
    elif access_mask & 0x120089 == 0x120089:  # Read
        permissions.append("Read")

    if access_mask & 0x100116 == 0x100116:  # Write
        if "Modify" not in permissions:
            permissions.append("Write")

    if not permissions:
        # Fall back to individual flags
        for flag, name in PERMISSION_FLAGS.items():
            if access_mask & flag:
                permissions.append(name)

    return ", ".join(permissions) if permissions else f"Special({hex(access_mask)})"


def get_sid_name(sid):
    """Convert SID to account name."""
    try:
        name, domain, _ = win32security.LookupAccountSid(None, sid)
        if domain:
            return f"{domain}\\{name}"
        return name
    except Exception:
        # Return SID string if lookup fails
        return win32security.ConvertSidToStringSid(sid)


def get_permissions(path, logger):
    """Get NTFS permissions for a file or folder."""
    permissions_list = []
    owner = ""

    try:
        # Get security descriptor
        sd = win32security.GetFileSecurity(
            path,
            win32security.OWNER_SECURITY_INFORMATION |
            win32security.DACL_SECURITY_INFORMATION
        )

        # Get owner
        owner_sid = sd.GetSecurityDescriptorOwner()
        if owner_sid:
            owner = get_sid_name(owner_sid)

        # Get DACL
        dacl = sd.GetSecurityDescriptorDacl()

        if dacl:
            for i in range(dacl.GetAceCount()):
                ace = dacl.GetAce(i)
                ace_type = ace[0][0]
                ace_flags = ace[0][1]
                access_mask = ace[1]
                sid = ace[2]

                # Get account name
                account_name = get_sid_name(sid)

                # Determine if inherited or explicit
                is_inherited = bool(ace_flags & win32security.INHERITED_ACE)
                inheritance_type = "Inherited" if is_inherited else "Explicit"

                # Determine allow or deny
                if ace_type == win32security.ACCESS_ALLOWED_ACE_TYPE:
                    access_type = "Allow"
                elif ace_type == win32security.ACCESS_DENIED_ACE_TYPE:
                    access_type = "Deny"
                else:
                    access_type = "Other"

                # Get permission string
                permission_str = get_access_mask_string(access_mask)

                permissions_list.append({
                    'account': account_name,
                    'access_type': access_type,
                    'permission': permission_str,
                    'inheritance': inheritance_type
                })

    except Exception as e:
        logger.debug(f"Error getting permissions for {path}: {str(e)}")
        raise

    return owner, permissions_list


def format_permissions(permissions_list):
    """Format permissions list into a single string for Excel cell."""
    if not permissions_list:
        return ""

    formatted = []
    for perm in permissions_list:
        entry = f"{perm['account']}:{perm['access_type']}:{perm['permission']}({perm['inheritance']})"
        formatted.append(entry)

    return "; ".join(formatted)


def get_file_info(path, is_dir):
    """Get file/folder information."""
    try:
        stat_info = os.stat(path)

        if is_dir:
            size = 0
            extension = ""
        else:
            size = stat_info.st_size
            extension = Path(path).suffix.lower()

        modified_time = datetime.fromtimestamp(stat_info.st_mtime)

        return {
            'size': size,
            'size_formatted': format_size(size) if not is_dir else "",
            'extension': extension,
            'modified': modified_time.strftime('%Y-%m-%d %H:%M:%S')
        }
    except Exception:
        return {
            'size': 0,
            'size_formatted': "",
            'extension': "",
            'modified': ""
        }


def setup_excel_workbook():
    """Create and setup Excel workbook with headers."""
    wb = Workbook()

    # Main data sheet
    ws_data = wb.active
    ws_data.title = "File_Folder_List"

    # Error sheet
    ws_errors = wb.create_sheet("Errors")

    # Define headers
    data_headers = [
        "Folder Path",
        "Name",
        "Type",
        "Extension",
        "Size (Bytes)",
        "Size (Formatted)",
        "Last Modified",
        "Owner",
        "Permissions"
    ]

    error_headers = [
        "Path",
        "Error Type",
        "Error Message",
        "Timestamp"
    ]

    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write data headers
    for col, header in enumerate(data_headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write error headers
    for col, header in enumerate(error_headers, 1):
        cell = ws_errors.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths for data sheet
    column_widths = [80, 40, 10, 10, 15, 15, 20, 30, 100]
    for col, width in enumerate(column_widths, 1):
        ws_data.column_dimensions[get_column_letter(col)].width = width

    # Set column widths for error sheet
    error_widths = [80, 20, 60, 20]
    for col, width in enumerate(error_widths, 1):
        ws_errors.column_dimensions[get_column_letter(col)].width = width

    # Freeze top row
    ws_data.freeze_panes = "A2"
    ws_errors.freeze_panes = "A2"

    # Enable auto filter
    ws_data.auto_filter.ref = f"A1:I1"
    ws_errors.auto_filter.ref = f"A1:D1"

    return wb, ws_data, ws_errors


def scan_directory(root_path, wb, ws_data, ws_errors, logger, exclusion_patterns=None):
    """Scan directory and write results to Excel.

    Args:
        root_path: Root directory to scan
        wb: Excel workbook
        ws_data: Data worksheet
        ws_errors: Errors worksheet
        logger: Logger instance
        exclusion_patterns: List of patterns to exclude (optional)
    """
    if exclusion_patterns is None:
        exclusion_patterns = []

    data_row = 2
    error_row = 2

    folder_count = 0
    file_count = 0
    error_count = 0
    excluded_count = 0

    logger.info(f"Starting scan of: {root_path}")
    if exclusion_patterns:
        logger.info(f"Exclusion patterns loaded: {len(exclusion_patterns)}")
    logger.info("-" * 60)

    # First, process the root folder itself
    try:
        owner, permissions = get_permissions(root_path, logger)
        file_info = get_file_info(root_path, True)

        # For root folder, show its parent directory as the folder path
        root_parent = os.path.dirname(root_path) or root_path
        ws_data.cell(row=data_row, column=1, value=root_parent)
        ws_data.cell(row=data_row, column=2, value=os.path.basename(root_path) or root_path)
        ws_data.cell(row=data_row, column=3, value="Folder")
        ws_data.cell(row=data_row, column=4, value="")
        ws_data.cell(row=data_row, column=5, value=0)
        ws_data.cell(row=data_row, column=6, value="")
        ws_data.cell(row=data_row, column=7, value=file_info['modified'])
        ws_data.cell(row=data_row, column=8, value=owner)
        ws_data.cell(row=data_row, column=9, value=format_permissions(permissions))

        data_row += 1
        folder_count += 1

    except Exception as e:
        ws_errors.cell(row=error_row, column=1, value=root_path)
        ws_errors.cell(row=error_row, column=2, value=type(e).__name__)
        ws_errors.cell(row=error_row, column=3, value=str(e))
        ws_errors.cell(row=error_row, column=4, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        error_row += 1
        error_count += 1
        logger.warning(f"Error accessing root path: {str(e)}")

    # Walk through directory tree
    for current_dir, subdirs, files in os.walk(root_path):
        # Filter out excluded subdirectories (modifying subdirs in-place skips them in os.walk)
        excluded_subdirs = []
        for subdir in subdirs[:]:  # Use slice copy to allow modification during iteration
            folder_path = os.path.join(current_dir, subdir)
            if is_path_excluded(folder_path, exclusion_patterns):
                subdirs.remove(subdir)  # This prevents os.walk from descending into this folder
                excluded_subdirs.append(subdir)
                excluded_count += 1
                logger.info(f"Excluded folder: {folder_path}")

        # Process non-excluded subdirectories
        for subdir in subdirs:
            folder_path = os.path.join(current_dir, subdir)
            folder_count += 1

            if folder_count % 100 == 0:
                logger.info(f"Progress: {folder_count} folders, {file_count} files scanned, {excluded_count} excluded...")

            try:
                owner, permissions = get_permissions(folder_path, logger)
                file_info = get_file_info(folder_path, True)

                # Folder path is the parent directory (current_dir)
                ws_data.cell(row=data_row, column=1, value=current_dir)
                ws_data.cell(row=data_row, column=2, value=subdir)
                ws_data.cell(row=data_row, column=3, value="Folder")
                ws_data.cell(row=data_row, column=4, value="")
                ws_data.cell(row=data_row, column=5, value=0)
                ws_data.cell(row=data_row, column=6, value="")
                ws_data.cell(row=data_row, column=7, value=file_info['modified'])
                ws_data.cell(row=data_row, column=8, value=owner)
                ws_data.cell(row=data_row, column=9, value=format_permissions(permissions))

                data_row += 1

            except Exception as e:
                ws_errors.cell(row=error_row, column=1, value=folder_path)
                ws_errors.cell(row=error_row, column=2, value=type(e).__name__)
                ws_errors.cell(row=error_row, column=3, value=str(e))
                ws_errors.cell(row=error_row, column=4, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                error_row += 1
                error_count += 1
                logger.debug(f"Error on folder {folder_path}: {str(e)}")

        # Process files
        for filename in files:
            file_path = os.path.join(current_dir, filename)
            file_count += 1

            if file_count % 500 == 0:
                logger.info(f"Progress: {folder_count} folders, {file_count} files scanned, {excluded_count} excluded...")

            try:
                owner, permissions = get_permissions(file_path, logger)
                file_info = get_file_info(file_path, False)

                # Folder path is the containing directory (current_dir)
                ws_data.cell(row=data_row, column=1, value=current_dir)
                ws_data.cell(row=data_row, column=2, value=filename)
                ws_data.cell(row=data_row, column=3, value="File")
                ws_data.cell(row=data_row, column=4, value=file_info['extension'])
                ws_data.cell(row=data_row, column=5, value=file_info['size'])
                ws_data.cell(row=data_row, column=6, value=file_info['size_formatted'])
                ws_data.cell(row=data_row, column=7, value=file_info['modified'])
                ws_data.cell(row=data_row, column=8, value=owner)
                ws_data.cell(row=data_row, column=9, value=format_permissions(permissions))

                data_row += 1

            except Exception as e:
                ws_errors.cell(row=error_row, column=1, value=file_path)
                ws_errors.cell(row=error_row, column=2, value=type(e).__name__)
                ws_errors.cell(row=error_row, column=3, value=str(e))
                ws_errors.cell(row=error_row, column=4, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                error_row += 1
                error_count += 1
                logger.debug(f"Error on file {file_path}: {str(e)}")

    return folder_count, file_count, error_count, excluded_count


def main():
    # Parse command line arguments
    parser = argparse.ArgumentParser(
        description='File Server Audit Tool - Scans folders and generates Excel report with permissions',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
    python file_server_audit.py "D:\\Data"
    python file_server_audit.py "D:\\Data" --exclude "D:\\Data\\Temp"
    python file_server_audit.py "D:\\Data" -x "*\\Archive" -x "*\\Backup"
    python file_server_audit.py "D:\\Data" --exclude-file exclusions.txt
    python file_server_audit.py "D:\\Data" -x "*\\Temp" --exclude-file exclusions.txt

Exclusion Options:
    -x, --exclude       Exclude a folder path pattern (can be used multiple times)
    -e, --exclude-file  Path to text file with exclusion patterns (one per line)

    Both options can be combined. Supports wildcards: * (any chars), ? (single char)
    Matching is case-insensitive.
        '''
    )
    parser.add_argument('root_path', help='Root path to scan (local path or UNC path)')
    parser.add_argument('--exclude', '-x', dest='exclude_patterns', action='append', default=[],
                        metavar='PATTERN', help='Folder path pattern to exclude (can be used multiple times)')
    parser.add_argument('--exclude-file', '-e', dest='exclude_file',
                        help='Path to text file containing folder paths to exclude (one per line)')

    args = parser.parse_args()
    root_path = args.root_path
    exclude_file = args.exclude_file
    cli_exclude_patterns = args.exclude_patterns

    # Validate root path
    if not os.path.exists(root_path):
        print(f"ERROR: Path does not exist: {root_path}")
        sys.exit(1)

    if not os.path.isdir(root_path):
        print(f"ERROR: Path is not a directory: {root_path}")
        sys.exit(1)

    # Setup output directory
    script_dir = Path(__file__).parent.absolute()
    output_dir = script_dir / "output"
    output_dir.mkdir(exist_ok=True)

    # Setup logging
    logger = setup_logging(output_dir)

    logger.info("=" * 60)
    logger.info("FILE SERVER AUDIT TOOL")
    logger.info("=" * 60)
    logger.info(f"Root Path: {root_path}")
    logger.info(f"Output Directory: {output_dir}")
    if cli_exclude_patterns:
        logger.info(f"CLI Exclusion Patterns: {len(cli_exclude_patterns)}")
    if exclude_file:
        logger.info(f"Exclusion File: {exclude_file}")
    logger.info(f"Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 60)

    # Load exclusion patterns from file
    exclusion_patterns = load_exclusion_patterns(exclude_file, logger)

    # Add CLI exclusion patterns
    if cli_exclude_patterns:
        for pattern in cli_exclude_patterns:
            # Normalize pattern (consistent separators, lowercase)
            normalized = pattern.replace('/', '\\').lower()
            exclusion_patterns.append(normalized)
            logger.debug(f"Added CLI exclusion pattern: {normalized}")
        logger.info(f"Added {len(cli_exclude_patterns)} CLI exclusion pattern(s)")

    # Log total exclusion patterns
    if exclusion_patterns:
        logger.info(f"Total exclusion patterns: {len(exclusion_patterns)}")

    # Setup Excel workbook
    wb, ws_data, ws_errors = setup_excel_workbook()

    # Scan directory
    start_time = datetime.now()
    folder_count, file_count, error_count, excluded_count = scan_directory(
        root_path, wb, ws_data, ws_errors, logger, exclusion_patterns
    )
    end_time = datetime.now()

    # Generate output filename
    safe_path_name = root_path.replace("\\", "_").replace(":", "").replace("/", "_")
    if len(safe_path_name) > 50:
        safe_path_name = safe_path_name[:50]

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = output_dir / f"FileAudit_{safe_path_name}_{timestamp}.xlsx"

    # Save workbook
    logger.info("-" * 60)
    logger.info("Saving Excel file...")
    wb.save(output_file)

    # Calculate duration
    duration = end_time - start_time

    # Summary
    logger.info("=" * 60)
    logger.info("SCAN COMPLETE")
    logger.info("=" * 60)
    logger.info(f"Folders Scanned: {folder_count:,}")
    logger.info(f"Files Scanned: {file_count:,}")
    logger.info(f"Total Items: {folder_count + file_count:,}")
    logger.info(f"Folders Excluded: {excluded_count:,}")
    logger.info(f"Errors: {error_count:,}")
    logger.info(f"Duration: {duration}")
    logger.info(f"Output File: {output_file}")
    logger.info("=" * 60)

    print(f"\nOutput saved to: {output_file}")


if __name__ == "__main__":
    main()
